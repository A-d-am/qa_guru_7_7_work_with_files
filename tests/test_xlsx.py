from openpyxl import load_workbook
import os
from conftest import RESOURCES_DIR


# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    workbook_path = os.path.join(RESOURCES_DIR,'file_example_XLSX_50.xlsx' )
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    row = 3
    column = 2
    value_in_given_cell = sheet.cell(row=row, column=column).value
    assert value_in_given_cell == 'Mara', (f'Wrong value in cell {row = } {column = }: '
                                           f'expected {"Mara"}, got {value_in_given_cell}')
